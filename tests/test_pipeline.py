from __future__ import annotations

"""主切分链路测试，覆盖解析、OCR 回退、边界增强和统一选择器。"""

from io import BytesIO

from openpyxl import Workbook
import xlwt

from app.core.config import settings
from app.models.schemas import ChunkOptions, DocumentNode
from app.services.boundary import BoundaryDecisionEngine
from app.services.llm import LlmBoundaryRefiner
from app.services.parser import PdfParser, get_parser
from app.services.pipeline import DocumentChunkPipeline
from app.services.selection import RuntimeSelector
from app.services.token_counter import TokenCounter


def test_xlsx_parser_extracts_each_sheet_as_table_node() -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "SheetA"
    first.append(["Field", "Value"])
    first.append(["Color", "Red"])
    second = workbook.create_sheet("SheetB")
    second.append(["Param", "Desc"])
    second.append(["Size", "Large"])
    stream = BytesIO()
    workbook.save(stream)

    parser = get_parser("demo.xlsx")
    nodes = parser.parse(stream.getvalue(), "demo.xlsx")

    assert len(nodes) == 4
    assert nodes[0].node_type == "title"
    assert nodes[1].node_type == "table"
    assert nodes[0].text == "SheetA"
    assert nodes[1].source_meta["sheet_name"] == "SheetA"
    assert "Field | Value" in nodes[1].text
    assert nodes[2].text == "SheetB"
    assert nodes[3].source_meta["sheet_name"] == "SheetB"


def test_xls_parser_extracts_sheet_as_table_node() -> None:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1")
    sheet.write(0, 0, "Field")
    sheet.write(0, 1, "Value")
    sheet.write(1, 0, "Color")
    sheet.write(1, 1, "Red")
    stream = BytesIO()
    workbook.save(stream)

    parser = get_parser("demo.xls")
    nodes = parser.parse(stream.getvalue(), "demo.xls")

    assert len(nodes) == 2
    assert nodes[0].node_type == "title"
    assert nodes[1].node_type == "table"
    assert nodes[0].text == "Sheet1"
    assert nodes[1].source_meta["sheet_name"] == "Sheet1"
    assert "Field | Value" in nodes[1].text


def test_pipeline_keeps_content_in_the_same_section() -> None:
    payload = (
        "# Chapter One\n\n"
        "Introduction to the first chapter.\n\n"
        "## Parameters\n"
        "| Field | Meaning |\n"
        "| --- | --- |\n"
        "| A | Value |\n\n"
        "# Chapter Two\n\n"
        "Explanation for the second chapter."
    )

    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(payload.encode("utf-8"), "demo.md")

    assert result.total_chunks >= 3
    section_paths = [tuple(chunk.section_path) for chunk in result.chunks]
    assert ("Chapter One",) in section_paths
    assert ("Chapter One", "Parameters") in section_paths
    assert ("Chapter Two",) in section_paths


def test_pipeline_binds_title_with_following_paragraph_chunk() -> None:
    payload = "# Product Guide\n\nThis paragraph should stay with the title."

    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(payload.encode("utf-8"), "demo.md")

    assert result.total_chunks == 1
    assert result.chunks[0].section_path == ["Product Guide"]
    assert result.chunks[0].metadata.chunk_type in {"paragraph", "mixed"}
    assert result.chunks[0].text.startswith("Product Guide")
    assert "This paragraph should stay with the title." in result.chunks[0].text


def test_pipeline_binds_title_with_following_list_chunk() -> None:
    payload = "# Checklist\n\n- item one\n- item two"

    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(payload.encode("utf-8"), "demo.md")

    assert result.total_chunks == 1
    assert result.chunks[0].section_path == ["Checklist"]
    assert result.chunks[0].text.startswith("Checklist")
    assert "item one" in result.chunks[0].text
    assert "item two" in result.chunks[0].text


def test_pipeline_binds_title_with_following_table_chunk() -> None:
    payload = (
        "# Parameters\n\n"
        "| Field | Meaning |\n"
        "| --- | --- |\n"
        "| A | Value |\n"
    )

    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(payload.encode("utf-8"), "demo.md")

    assert result.total_chunks == 1
    assert result.chunks[0].section_path == ["Parameters"]
    assert result.chunks[0].text.startswith("Parameters")
    assert "Field | Meaning" in result.chunks[0].text


def test_pipeline_splits_oversized_text_without_cutting_everything_to_one_chunk() -> None:
    large_paragraph = "This is a very long paragraph. " * ((settings.max_chunk_tokens * 3) + 40)
    payload = f"# Large Document\n\n{large_paragraph}"

    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(payload.encode("utf-8"), "large.md")

    assert result.total_chunks >= 2
    assert all(chunk.text for chunk in result.chunks)


def test_pipeline_llm_refiner_can_merge_adjacent_blocks() -> None:
    payload = (
        "# Product Guide\n\n"
        "This is a short explanation.\n"
        "- This bullet is strongly related to the previous line."
    )

    original_merge = LlmBoundaryRefiner.decide_merge
    original_flag = settings.llm_enabled

    def always_merge(self: LlmBoundaryRefiner, left_text: str, right_text: str, options=None) -> bool:
        return True

    settings.llm_enabled = True
    LlmBoundaryRefiner.decide_merge = always_merge
    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(
        payload.encode("utf-8"),
        "merge.md",
        ChunkOptions(min_chunk_tokens=20, target_chunk_tokens=80, max_chunk_tokens=160),
    )
    LlmBoundaryRefiner.decide_merge = original_merge
    settings.llm_enabled = original_flag

    assert result.total_chunks == 1
    merged_chunk = result.chunks[0]
    assert merged_chunk.text.startswith("Product Guide")
    assert "short explanation" in merged_chunk.text
    assert "strongly related" in merged_chunk.text


def test_pipeline_llm_refiner_falls_back_to_rule_blocks_on_error() -> None:
    payload = (
        "# Section One\n"
        "First section body.\n"
        "# Section Two\n"
        "Second section body."
    )

    original_merge = LlmBoundaryRefiner.decide_merge
    original_flag = settings.llm_enabled

    def broken(self: LlmBoundaryRefiner, left_text: str, right_text: str, options=None) -> bool:
        raise RuntimeError("llm failure")

    settings.llm_enabled = True
    LlmBoundaryRefiner.decide_merge = broken
    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(
        payload.encode("utf-8"),
        "fallback.md",
        ChunkOptions(min_chunk_tokens=20, target_chunk_tokens=80, max_chunk_tokens=160),
    )
    LlmBoundaryRefiner.decide_merge = original_merge
    settings.llm_enabled = original_flag

    assert result.total_chunks >= 2
    section_paths = [tuple(chunk.section_path) for chunk in result.chunks]
    assert ("Section One",) in section_paths
    assert ("Section Two",) in section_paths


def test_pipeline_uses_visual_analyzer_for_image_documents() -> None:
    original = DocumentChunkPipeline._analyze_image_document

    def fake_analyze(self: DocumentChunkPipeline, file_bytes: bytes, filename: str, options) -> list[DocumentNode]:
        assert filename == "demo.png"
        return [
            DocumentNode(
                node_id="img-node",
                node_type="paragraph",
                text="The image contains product instructions and precautions.",
                source_meta={"modality": "vision"},
            )
        ]

    DocumentChunkPipeline._analyze_image_document = fake_analyze
    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(b"fake-image", "demo.png")
    DocumentChunkPipeline._analyze_image_document = original

    assert result.total_chunks == 1
    assert "product instructions" in result.chunks[0].text
    assert result.chunks[0].metadata.chunk_type == "paragraph"


def test_pipeline_falls_back_to_vision_ocr_for_scanned_pdf() -> None:
    original_parser = DocumentChunkPipeline._parse_document
    original_ocr = DocumentChunkPipeline._analyze_pdf_with_vision

    def fake_parse(self: DocumentChunkPipeline, file_bytes: bytes, filename: str) -> list[DocumentNode]:
        return []

    def fake_ocr(self: DocumentChunkPipeline, file_bytes: bytes, filename: str, options) -> list[DocumentNode]:
        return [
            DocumentNode(
                node_id="pdf-ocr-node",
                node_type="paragraph",
                text="OCR result extracted from an image-only PDF.",
                source_page=1,
                source_meta={"modality": "vision_ocr"},
            )
        ]

    DocumentChunkPipeline._parse_document = fake_parse
    DocumentChunkPipeline._analyze_pdf_with_vision = fake_ocr
    pipeline = DocumentChunkPipeline()
    result = pipeline.chunk_bytes(b"%PDF-1.4 scanned", "scan.pdf")
    DocumentChunkPipeline._parse_document = original_parser
    DocumentChunkPipeline._analyze_pdf_with_vision = original_ocr

    assert result.total_chunks == 1
    assert "OCR result" in result.chunks[0].text
    assert result.chunks[0].metadata.page_no == [1]


def test_pdf_parser_removes_repeated_headers_and_footers() -> None:
    parser = PdfParser()
    nodes = [
        DocumentNode(node_id="1", node_type="paragraph", text="Internal Only", source_page=1),
        DocumentNode(node_id="2", node_type="paragraph", text="Chapter 1 Overview", source_page=1),
        DocumentNode(node_id="3", node_type="paragraph", text="Body content on page one.", source_page=1),
        DocumentNode(node_id="4", node_type="paragraph", text="Page 1", source_page=1),
        DocumentNode(node_id="5", node_type="paragraph", text="Internal Only", source_page=2),
        DocumentNode(node_id="6", node_type="paragraph", text="Body content on page two.", source_page=2),
        DocumentNode(node_id="7", node_type="paragraph", text="Page 2", source_page=2),
    ]

    cleaned = parser._remove_repeated_page_noise(nodes)

    texts = [node.text for node in cleaned]
    assert "Internal Only" not in texts
    assert "Page 1" not in texts
    assert "Page 2" not in texts
    assert "Body content on page one." in texts
    assert "Body content on page two." in texts


def test_boundary_engine_merges_when_similarity_is_high() -> None:
    engine = BoundaryDecisionEngine()
    left = [DocumentNode(node_id="1", node_type="paragraph", text="Product usage guide", source_meta={"section_path": ["Guide"]})]
    right = [DocumentNode(node_id="2", node_type="paragraph", text="Usage details and precautions", source_meta={"section_path": ["Guide"]})]

    original_score = engine.similarity_scorer.score
    engine.similarity_scorer.score = lambda a, b, options=None: 0.95
    decision = engine.should_merge(left, right, ChunkOptions(max_chunk_tokens=160, min_chunk_tokens=20))
    engine.similarity_scorer.score = original_score

    assert decision["merge"] is True
    assert decision["strategy"] == "similarity_high"
    assert decision["similarity_score"] == 0.95


def test_boundary_engine_keeps_when_similarity_is_low() -> None:
    engine = BoundaryDecisionEngine()
    left = [DocumentNode(node_id="1", node_type="paragraph", text="Product usage guide", source_meta={"section_path": ["Guide"]})]
    right = [DocumentNode(node_id="2", node_type="paragraph", text="Refund policy and billing", source_meta={"section_path": ["Guide"]})]

    original_score = engine.similarity_scorer.score
    engine.similarity_scorer.score = lambda a, b, options=None: 0.41
    decision = engine.should_merge(left, right, ChunkOptions(max_chunk_tokens=160, min_chunk_tokens=20))
    engine.similarity_scorer.score = original_score

    assert decision["merge"] is False
    assert decision["strategy"] == "similarity_low"
    assert decision["similarity_score"] == 0.41


def test_boundary_engine_uses_llm_for_gray_zone() -> None:
    original_flag = settings.llm_enabled
    settings.llm_enabled = True
    engine = BoundaryDecisionEngine()
    left = [DocumentNode(node_id="1", node_type="paragraph", text="Product usage guide", source_meta={"section_path": ["Guide"]})]
    right = [DocumentNode(node_id="2", node_type="paragraph", text="Usage detail follows", source_meta={"section_path": ["Guide"]})]

    original_score = engine.similarity_scorer.score
    original_merge = engine.llm_refiner.decide_merge
    engine.similarity_scorer.score = lambda a, b, options=None: 0.8
    engine.llm_refiner.decide_merge = lambda a, b, options=None: True
    decision = engine.should_merge(left, right, ChunkOptions(max_chunk_tokens=160, min_chunk_tokens=20))
    engine.similarity_scorer.score = original_score
    engine.llm_refiner.decide_merge = original_merge
    settings.llm_enabled = original_flag

    assert decision["merge"] is True
    assert decision["strategy"] == "llm_gray"
    assert decision["similarity_score"] == 0.8


def test_boundary_engine_falls_back_when_similarity_service_fails() -> None:
    original_flag = settings.llm_enabled
    settings.llm_enabled = True
    engine = BoundaryDecisionEngine()
    left = [DocumentNode(node_id="1", node_type="paragraph", text="Product usage guide", source_meta={"section_path": ["Guide"]})]
    right = [DocumentNode(node_id="2", node_type="paragraph", text="Usage detail follows", source_meta={"section_path": ["Guide"]})]

    original_score = engine.similarity_scorer.score
    original_merge = engine.llm_refiner.decide_merge

    def broken_score(a: str, b: str, options=None) -> float:
        raise RuntimeError("embedding unavailable")

    engine.similarity_scorer.score = broken_score
    engine.llm_refiner.decide_merge = lambda a, b, options=None: False
    decision = engine.should_merge(left, right, ChunkOptions(max_chunk_tokens=160, min_chunk_tokens=20))
    engine.similarity_scorer.score = original_score
    engine.llm_refiner.decide_merge = original_merge
    settings.llm_enabled = original_flag

    assert decision["merge"] is False
    assert decision["strategy"] == "llm_fallback"
    assert decision["similarity_score"] is None


def test_runtime_selector_uses_service_side_models() -> None:
    selector = RuntimeSelector()
    selection = selector.resolve(ChunkOptions())

    assert selection.text_model == settings.text_model
    assert selection.flash_model == (settings.flash_model or settings.text_model)
    assert selection.vision_model == settings.vision_model
    assert selection.embedding_base_url == settings.embedding_base_url
    assert selection.embedding_model == settings.embedding_model
    assert selection.embedding_api_key == settings.embedding_api_key


def test_token_counter_uses_cache_for_duplicate_text() -> None:
    counter = TokenCounter()
    text = "repeat this text"
    first = counter.count(text)
    second = counter.count(text)

    assert first == second
